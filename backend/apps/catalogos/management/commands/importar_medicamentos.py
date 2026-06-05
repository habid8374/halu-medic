"""
Importa el catálogo de medicamentos CUM desde el Superhomologador MinSalud.

Fuente: hoja 'CUM' del archivo SUPERHOMOLOGADOR2026BASICO_CON_COD_REPS.xlsm
Columnas usadas:
  principioactivo, expediente (= CUM), cantidad, unidadmedida,
  unidadreferencia, formafarmaceutica, registrosanitario,
  fechavencimiento, estadoregistro

Uso:
  python manage.py importar_medicamentos /ruta/SUPERHOMOLOGADOR2026.xlsm
  python manage.py importar_medicamentos /ruta/archivo.xlsm --hoja CUM
  python manage.py importar_medicamentos /ruta/archivo.xlsm --solo-vigentes
"""
from django.core.management.base import BaseCommand, CommandError
from django.db import connection

import os


class Command(BaseCommand):
    help = 'Importa catálogo CUM de medicamentos desde el Superhomologador MinSalud'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str, help='Ruta al archivo .xlsm/.xlsx')
        parser.add_argument('--hoja', default='CUM', help='Nombre de la hoja (default: CUM)')
        parser.add_argument('--solo-vigentes', action='store_true',
                            help='Solo importa registros con estadoregistro=Vigente')
        parser.add_argument('--actualizar', action='store_true',
                            help='Actualiza registros existentes (default: skip duplicados)')

    def handle(self, *args, **options):
        from apps.catalogos.models import CatalogoMedicamento
        import openpyxl

        ruta = options['archivo']
        if not os.path.exists(ruta):
            raise CommandError(f'Archivo no encontrado: {ruta}')

        self.stdout.write(f'Cargando {ruta}...')
        try:
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f'Error al abrir el archivo: {e}')

        hoja_nombre = options['hoja']
        if hoja_nombre not in wb.sheetnames:
            raise CommandError(
                f'Hoja "{hoja_nombre}" no encontrada. Hojas disponibles: {wb.sheetnames}'
            )

        ws = wb[hoja_nombre]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError('La hoja está vacía')

        # Detectar columnas por cabecera
        cabecera = [str(c).strip().lower() if c else '' for c in rows[0]]
        col = {name: i for i, name in enumerate(cabecera)}

        required = ['principioactivo', 'expediente']
        for r in required:
            if r not in col:
                raise CommandError(
                    f'Columna requerida "{r}" no encontrada. Cabeceras: {cabecera}'
                )

        def get(row, name, default=''):
            idx = col.get(name)
            if idx is None:
                return default
            val = row[idx]
            return str(val).strip() if val is not None else default

        solo_vigentes = options['solo_vigentes']
        actualizar    = options['actualizar']

        creados = actualizados = omitidos = 0
        batch = []
        BATCH_SIZE = 500

        for i, row in enumerate(rows[1:], start=2):
            cum = get(row, 'expediente')
            if not cum or cum == 'None':
                continue

            estado = get(row, 'estadoregistro', 'Vigente')
            vigente = 'vigente' in estado.lower()

            if solo_vigentes and not vigente:
                omitidos += 1
                continue

            principio = get(row, 'principioactivo')
            if not principio or principio == 'None':
                continue

            # Concentración: cantidad + unidad + referencia
            cantidad    = get(row, 'cantidad')
            unidad      = get(row, 'unidadmedida')
            referencia  = get(row, 'unidadreferencia')
            partes = [p for p in [cantidad, unidad, referencia] if p and p != 'None']
            concentracion = ' '.join(partes)[:150]

            forma   = get(row, 'formafarmaceutica')[:150]
            reg_san = get(row, 'registrosanitario')[:60]

            batch.append(CatalogoMedicamento(
                cum=cum[:20],
                principio_activo=principio[:400],
                concentracion=concentracion,
                forma_farmaceutica=forma,
                registro_sanitario=reg_san,
                vigente=vigente,
            ))

            if len(batch) >= BATCH_SIZE:
                self._guardar_batch(batch, actualizar)
                creados += len(batch)
                batch = []
                self.stdout.write(f'  Procesadas {i} filas...', ending='\r')
                self.stdout.flush()

        if batch:
            self._guardar_batch(batch, actualizar)
            creados += len(batch)

        wb.close()
        self.stdout.write(self.style.SUCCESS(
            f'\nListo: {creados} registros importados, {omitidos} omitidos.'
        ))

    def _guardar_batch(self, batch, actualizar):
        from apps.catalogos.models import CatalogoMedicamento
        if actualizar:
            CatalogoMedicamento.objects.bulk_create(
                batch,
                update_conflicts=True,
                update_fields=['principio_activo', 'concentracion',
                               'forma_farmaceutica', 'registro_sanitario', 'vigente'],
                unique_fields=['cum'],
            )
        else:
            CatalogoMedicamento.objects.bulk_create(batch, ignore_conflicts=True)
