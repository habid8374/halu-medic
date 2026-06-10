"""
API para manuales tarifarios.
  /api/tarifas/                    - CRUD de manuales
  /api/tarifas/{id}/items/         - ítems del manual
  /api/tarifas/{id}/importar/      - importar Excel/CSV
  /api/tarifas/precio/             - buscar precio CUPS para un paciente
"""
import io
from decimal import Decimal
from rest_framework import serializers, viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.tarifas.models import ManualTarifario, ItemTarifario


class ItemTarifarioSerializer(serializers.ModelSerializer):
    valor_final = serializers.DecimalField(max_digits=14, decimal_places=2, read_only=True)

    class Meta:
        model = ItemTarifario
        fields = ['id', 'cups', 'descripcion', 'valor_base', 'valor_final', 'es_paquete', 'cups_rips']


class ManualTarifarioSerializer(serializers.ModelSerializer):
    total_items = serializers.SerializerMethodField()

    class Meta:
        model = ManualTarifario
        fields = [
            'id', 'nombre', 'tipo', 'porcentaje_ajuste',
            'es_predeterminado', 'activo', 'vigente_desde',
            'vigente_hasta', 'observaciones', 'creado_en', 'total_items',
        ]
        read_only_fields = ['id', 'creado_en']

    def get_total_items(self, obj):
        return obj.items.count()


class ManualTarifarioViewSet(viewsets.ModelViewSet):
    serializer_class = ManualTarifarioSerializer
    ordering = ['nombre']

    def get_queryset(self):
        return ManualTarifario.objects.all()

    # ── Ítems del manual ──────────────────────────────────────────────────────

    @action(detail=True, methods=['get'], url_path='items')
    def listar_items(self, request, pk=None):
        manual = self.get_object()
        q = request.query_params.get('search')
        qs = manual.items.all()
        if q:
            from django.db.models import Q
            qs = qs.filter(Q(cups__icontains=q) | Q(descripcion__icontains=q))
        page = self.paginate_queryset(qs)
        if page is not None:
            return self.get_paginated_response(ItemTarifarioSerializer(page, many=True).data)
        return Response(ItemTarifarioSerializer(qs, many=True).data)

    @action(detail=True, methods=['post'], url_path='items/agregar')
    def agregar_item(self, request, pk=None):
        manual = self.get_object()
        s = ItemTarifarioSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        item = s.save(manual=manual)
        return Response(ItemTarifarioSerializer(item).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['patch'], url_path='items/(?P<item_id>[^/.]+)/editar')
    def editar_item(self, request, pk=None, item_id=None):
        item = ItemTarifario.objects.get(pk=item_id, manual=self.get_object())
        s = ItemTarifarioSerializer(item, data=request.data, partial=True)
        s.is_valid(raise_exception=True)
        return Response(ItemTarifarioSerializer(s.save()).data)

    @action(detail=True, methods=['delete'], url_path='items/(?P<item_id>[^/.]+)/eliminar')
    def eliminar_item(self, request, pk=None, item_id=None):
        ItemTarifario.objects.filter(pk=item_id, manual=self.get_object()).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # ── Importar desde Excel/CSV ──────────────────────────────────────────────

    @action(detail=True, methods=['post'], url_path='importar')
    def importar(self, request, pk=None):
        """
        POST multipart/form-data con campo 'archivo' (.xlsx, .xls, .csv).
        Columnas detectadas automáticamente: cups/codigo, descripcion/nombre, valor/precio.
        Devuelve: { importados, actualizados, errores }
        """
        manual = self.get_object()
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'error': 'Debes enviar el campo "archivo".'}, status=400)

        nombre = archivo.name.lower()
        try:
            if nombre.endswith('.csv'):
                filas = _parse_csv(archivo)
            else:
                filas = _parse_excel(archivo)
        except ValueError as e:
            return Response({'error': str(e)}, status=400)
        except Exception as e:
            return Response({'error': f'No se pudo leer el archivo ({type(e).__name__}): {e}'}, status=400)

        # Parsear y validar filas
        objetos = []
        errores = 0
        mensajes_error = []
        vistos = set()
        for fila in filas:
            cups = fila.get('cups', '').strip()
            if not cups or cups in vistos:
                if not cups:
                    errores += 1
                continue
            vistos.add(cups)
            try:
                valor_raw = str(fila.get('valor', 0) or 0).replace(',', '.').replace(' ', '').replace('$', '')
                valor = Decimal(valor_raw) if valor_raw else Decimal('0')
            except Exception as e:
                errores += 1
                mensajes_error.append(f'CUPS {cups}: valor inválido — {e}')
                continue
            desc = str(fila.get('descripcion', '') or '')[:400]
            objetos.append(ItemTarifario(
                manual=manual, cups=cups, descripcion=desc,
                valor_base=valor, es_paquete=False, cups_rips='',
            ))

        # Bulk upsert en lotes de 500
        BATCH = 500
        importados = actualizados = 0
        existentes = set(manual.items.values_list('cups', flat=True))
        nuevos   = [o for o in objetos if o.cups not in existentes]
        updates  = [o for o in objetos if o.cups in existentes]

        for i in range(0, len(nuevos), BATCH):
            creados = ItemTarifario.objects.bulk_create(nuevos[i:i+BATCH], ignore_conflicts=False)
            importados += len(creados)

        for i in range(0, len(updates), BATCH):
            ItemTarifario.objects.bulk_create(
                updates[i:i+BATCH],
                update_conflicts=True,
                unique_fields=['manual', 'cups'],
                update_fields=['descripcion', 'valor_base'],
            )
            actualizados += len(updates[i:i+BATCH])

        return Response({
            'importados': importados,
            'actualizados': actualizados,
            'errores': errores,
            'total_items': manual.items.count(),
            'detalle_errores': mensajes_error[:20],
        })

    # ── Plantilla CSV para importar ítems ────────────────────────────────────

    @action(detail=False, methods=['get'], url_path='plantilla')
    def plantilla(self, request):
        """Descarga plantilla Excel lista para llenar e importar."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Tarifario'

        # Estilos
        hdr_font  = Font(bold=True, color='FFFFFF', size=11)
        hdr_fill  = PatternFill('solid', fgColor='0F2D5E')
        hdr_align = Alignment(horizontal='center', vertical='center')
        ex_fill   = PatternFill('solid', fgColor='EFF6FF')
        thin      = Side(style='thin', color='CCCCCC')
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ['CUPS', 'Descripcion', 'Valor']
        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = border

        ejemplos = [
            ('890201', 'CONSULTA DE PRIMERA VEZ POR MEDICINA GENERAL',          50000),
            ('890202', 'CONSULTA DE CONTROL O SEGUIMIENTO POR MEDICINA GENERAL', 45000),
            ('890301', 'CONSULTA DE URGENCIAS POR MEDICINA GENERAL',             65000),
            ('890406', 'CONSULTA DE CONTROL POR MEDICINA ESPECIALIZADA',         55000),
            ('841000', 'ELECTROCARDIOGRAMA',                                     40000),
            ('841100', 'ESPIROMETRIA SIMPLE',                                    38000),
            ('874000', 'RADIOGRAFIA DE TORAX ANTEROPOSTERIOR',                   35000),
            ('874100', 'ECOGRAFIA ABDOMINAL TOTAL',                              80000),
            ('874200', 'MAMOGRAFIA BILATERAL',                                   90000),
            ('903803', 'ALBUMINA EN SUERO',                                      18000),
            ('904210', 'GLUCOSA EN AYUNAS',                                      12000),
            ('904214', 'CREATININA EN SUERO',                                    15000),
            ('904200', 'HEMOGRAMA TIPO IV (CUADRO HEMATICO)',                    20000),
            ('904221', 'PARCIAL DE ORINA (UROANÁLISIS)',                         14000),
            ('890701', 'ESTANCIA HOSPITALARIA - MEDICINA GENERAL (DIA CAMA)',    85000),
            ('890703', 'ESTANCIA UCI ADULTOS',                                  350000),
        ]
        for i, fila in enumerate(ejemplos, 2):
            ws.append(list(fila))
            for col in range(1, 4):
                cell = ws.cell(row=i, column=col)
                cell.fill = ex_fill
                cell.border = border
                cell.alignment = Alignment(vertical='center')

        # Anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 16
        ws.row_dimensions[1].height = 22

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="plantilla_tarifario.xlsx"'
        resp['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return resp

    # ── Lookup de precio para un CUPS + paciente ──────────────────────────────

    @action(detail=False, methods=['get'], url_path='precio')
    def precio(self, request):
        """
        GET /api/tarifas/precio/?cups=890201&paciente=<uuid>
        Devuelve el valor_final del CUPS en la tarifa del paciente
        (o la predeterminada si no tiene asignada).
        """
        cups_code = request.query_params.get('cups')
        paciente_id = request.query_params.get('paciente')
        if not cups_code:
            return Response({'error': 'Parámetro cups requerido.'}, status=400)

        manual = None
        if paciente_id:
            try:
                from apps.pacientes.models import Paciente
                p = Paciente.objects.select_related('tarifa').get(pk=paciente_id)
                manual = p.tarifa
            except Paciente.DoesNotExist:
                pass

        if manual is None:
            manual = ManualTarifario.objects.filter(es_predeterminado=True, activo=True).first()

        if manual is None:
            return Response({'encontrado': False, 'valor': None, 'manual': None})

        try:
            item = manual.items.get(cups=cups_code)
            return Response({
                'encontrado': True,
                'valor': float(item.valor_final),
                'valor_base': float(item.valor_base),
                'manual_id': str(manual.id),
                'manual_nombre': manual.nombre,
                'porcentaje_ajuste': float(manual.porcentaje_ajuste),
            })
        except ItemTarifario.DoesNotExist:
            return Response({
                'encontrado': False,
                'valor': None,
                'manual_id': str(manual.id),
                'manual_nombre': manual.nombre,
            })


# ── Helpers de parseo ─────────────────────────────────────────────────────────

_CUPS_COLS  = {'cups', 'codigo', 'cod', 'code', 'cod_cups', 'codigo_cups', 'codigocups', 'cup'}
_DESC_COLS  = {'descripcion', 'nombre', 'description', 'name', 'procedimiento', 'desc', 'detalle', 'servicio'}
_VALOR_COLS = {'valor', 'precio', 'value', 'price', 'tarifa', 'valor_base', 'valorbase', 'costo', 'importe', 'vr'}


def _norm_header(h):
    import unicodedata
    s = unicodedata.normalize('NFD', str(h))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')  # elimina tildes
    return s.lower().strip().replace(' ', '_').replace('.', '').replace(':', '').replace('$', '').replace('(', '').replace(')', '')


def _map_headers(headers):
    """Retorna dict {cups, descripcion, valor} con los índices de columna."""
    m = {}
    for i, h in enumerate(headers):
        hn = _norm_header(h)
        if hn in _CUPS_COLS and 'cups' not in m:
            m['cups'] = i
        elif hn in _DESC_COLS and 'descripcion' not in m:
            m['descripcion'] = i
        elif hn in _VALOR_COLS and 'valor' not in m:
            m['valor'] = i
    return m


def _parse_excel(archivo):
    import openpyxl
    content = archivo.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c) if c is not None else '' for c in rows[0]]
    m = _map_headers(headers)
    if 'cups' not in m or 'valor' not in m:
        norm = [_norm_header(h) for h in headers]
        raise ValueError(
            f'Columnas no reconocidas. '
            f'Se detectaron: {headers} (normalizado: {norm}). '
            f'La primera fila debe tener columnas llamadas "CUPS" y "Valor" (o similar).'
        )
    result = []
    for row in rows[1:]:
        def _v(idx):
            return row[idx] if 0 <= idx < len(row) else None
        desc_idx = m.get('descripcion')
        result.append({
            'cups': str(_v(m['cups']) or '').strip(),
            'descripcion': str(_v(desc_idx) or '').strip() if desc_idx is not None else '',
            'valor': _v(m['valor']) or 0,
        })
    return result


def _parse_csv(archivo):
    import csv
    content = archivo.read().decode('utf-8-sig', errors='replace')
    reader = csv.DictReader(io.StringIO(content))
    m = _map_headers(reader.fieldnames or [])
    if 'cups' not in m or 'valor' not in m:
        raise ValueError('No se encontraron columnas CUPS y valor en el CSV.')
    result = []
    for row in reader:
        headers = list(row.keys())
        desc_key = headers[m['descripcion']] if 'descripcion' in m else None
        result.append({
            'cups': str(row.get(headers[m['cups']], '') or '').strip(),
            'descripcion': str(row.get(desc_key, '') or '').strip() if desc_key else '',
            'valor': row.get(headers[m['valor']], 0) or 0,
        })
    return result
